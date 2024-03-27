import tkinter as tk
from tkinter import ttk
import serial
import openpyxl
from tkinter import filedialog
import threading
import time
from tkinter import simpledialog
import re
import matplotlib.pyplot as plt

# Controlo de Sistema is portuguese for System Control
# Controlo Manual e Automático is portuguese for Manual and Automatic System
class StepperControlGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Controlo de Sistema")

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True)

        self.page = tk.Frame(self.notebook)

        self.notebook.add(self.page, text="Controlo Manual e Automático")
        self.create_page()
        
        # To adjust the window's close settings.
        self.root.protocol("WM_DELETE_WINDOW", self.on_window_close)

    def create_page(self):
        self.control_page = StepperControlPage(self.page, self.root)  # Turn root as argument

    def show_page(self):
        self.notebook.select(0)  # Automatic Control Index

    def on_window_close(self):
        #Recall to close the serial port using the close method of the actual page.
        self.control_page.close()
        # After, closes the actual page
        self.root.destroy()

class StepperControlPage:
    def __init__(self, parent, root):
        self.parent = parent
        self.root = root
        self.chosen_directory = ""
        # Variables to store force and momentum values
        self.forca_value = None
        self.momento_value = None

        # Variable of control of standard values
        self.default_tstep = tk.StringVar()
        self.default_tstep.set("5000")  # Standard value definition

        self.default_rep = tk.StringVar()
        self.default_rep.set("1")  # Standard value definition

        self.default_tmeas = tk.StringVar()
        self.default_tmeas.set("2000")  # Standard value definition

        self.default_ncount = tk.StringVar()
        self.default_ncount.set("4")  # Standard value definition

        # Configuration of serial communication for stepper motor control (adjust COM port as needed)
        self.stepper_serial_port = serial.Serial('COM6', 9600)

        # Create a frame called "input" to group the input fields
        self.input_frame = tk.Frame(self.parent)
        self.input_frame.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
        
        self.tstep_label = tk.Label(self.input_frame, text="TStep")
        self.tstep_label.grid(row=0, column=0, padx=2, pady=2, sticky=tk.W)
        self.tstep_entry = tk.Entry(self.input_frame, textvariable=self.default_tstep)
        self.tstep_entry.grid(row=0, column=0, padx=2, pady=2)  # Column 1

        self.rep_label = tk.Label(self.input_frame, text="Nº Rep")
        self.rep_label.grid(row=0, column=1, padx=2, pady=2, sticky=tk.W)  # Column 2
        self.rep_entry = tk.Entry(self.input_frame, textvariable=self.default_rep)
        self.rep_entry.grid(row=0, column=1, padx=2, pady=2)  # Column 3

        self.tmeas_label = tk.Label(self.input_frame, text="TMeas")
        self.tmeas_label.grid(row=0, column=2, padx=2, pady=2, sticky=tk.W)  # Column 4
        self.tmeas_entry = tk.Entry(self.input_frame, textvariable=self.default_tmeas)
        self.tmeas_entry.grid(row=0, column=3, padx=2, pady=2)  # Coluna 5

        self.ncount_label = tk.Label(self.input_frame, text="NºMeas")
        self.ncount_label.grid(row=0, column=4, padx=2, pady=2, sticky=tk.W)  # Column 6
        self.ncount_entry = tk.Entry(self.input_frame, textvariable=self.default_ncount)
        self.ncount_entry.grid(row=0, column=5, padx=2, pady=2)  # Column 7


        self.manual_label = tk.Label(self.input_frame, text="-------------------------------Controlo Manual---------------------------------------")
        self.manual_label.grid(row=2, column=0, padx=10, pady=10, columnspan=2, sticky="ew")

        # Input field for stepper motor 1
        #Inserir - ou + sempre
        self.stepper1_label = tk.Label(self.input_frame, text="Motor Linear [mm]")
        self.stepper1_label.grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        self.stepper1_entry = tk.Entry(self.input_frame)
        self.stepper1_entry.grid(row=4, column=0, padx=5, pady=5)

        # Input field for stepper motor 2
        self.stepper2_label = tk.Label(self.input_frame, text="Motor Angular [º]")
        self.stepper2_label.grid(row=3, column=1, padx=5, pady=5, sticky=tk.W)
        self.stepper2_entry = tk.Entry(self.input_frame)
        self.stepper2_entry.grid(row=4, column=1, padx=5, pady=5)

        # Button to send the values of the stepper motor 1 to the Arduino
        self.send_stepper1_button = tk.Button(self.input_frame, text="Mover Motor Linear", command=self.send_stepper1_values)
        self.send_stepper1_button.grid(row=5, column=0, padx=5, pady=5)
        
        # Button to send the values of the stepper motor 2 to the Arduino
        self.send_stepper2_button = tk.Button(self.input_frame, text="Mover Motor Angular", command=self.send_stepper2_values)
        self.send_stepper2_button.grid(row=5, column=1, padx=5, pady=5)


        # Create a frame called "output" to group the Force and Moment output fields
        self.output_frame = tk.Frame(self.parent)
        self.output_frame.grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
        
        # Text field to display the output
        self.output_text = tk.Text(self.output_frame, height=1, width=30, wrap="none")  # Set the height
        self.output_text.grid(row=4, column=0, padx=5, pady=5)
        self.output_text2 = tk.Text(self.output_frame, height=1, width=30, wrap="none")  # Set the height
        self.output_text2.grid(row=4, column=1, padx=5, pady=5)

        # Button to update the output field
        self.update_button = tk.Label(self.output_frame, text="Força [N]")
        self.update_button.grid(row=3, column=0, padx=5, pady=5)

        self.update2_button = tk.Label(self.output_frame, text="Momento [N.m]")
        self.update2_button.grid(row=3, column=1, padx=5, pady=5)

        
        # Button to choose save directory
        self.choose_dir_button = tk.Button(self.output_frame, text="Escolher Diretório", command=self.choose_directory)
        self.choose_dir_button.grid(row=5, column=0, padx=5, pady=5)

        # Button to save the data to file
        self.save_data_button = tk.Button(self.output_frame, text="Guardar Dados", command=self.save_data)
        self.save_data_button.grid(row=5, column=1, padx=5, pady=5)

        # Button to move the Linear Motor to the starting point
        self.move_to_initial1_button = tk.Button(self.output_frame, text="Home Motor Linear", command=self.move_to_initial1)
        self.move_to_initial1_button.grid(row=6, column=0, padx=5, pady=5)

        # Button to move the Angular Motor to the starting point
        self.move_to_initial2_button = tk.Button(self.output_frame, text="Home Motor Angular", command=self.move_to_initial2)
        self.move_to_initial2_button.grid(row=6, column=1, padx=5, pady=5)

        # Create a board called auto_sub_frame to group the AutoControl fields and buttons
        self.auto_sub_frame = tk.Frame(self.parent)
        self.auto_sub_frame.grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)

        self.auto_label = tk.Label(self.auto_sub_frame, text="----------------------------------------Controlo Automático------------------------------------------------")
        self.auto_label.grid(row=7, column=0, padx=10, pady=10, columnspan=2, sticky="ew")
        
        self.auto_text = tk.Text(self.auto_sub_frame, height=10, width=40, wrap=tk.WORD)
        self.auto_text.grid(row=8, column=0, padx=10, pady=10)
        self.auto_text.insert("1.0", "MA +23,3 ML -12,4\nMA -12,2 ML +10\nMA -12,3 ML +14")

        start_auto_button = tk.Button(self.auto_sub_frame, text="INICIAR", command=self.start_auto_control)
        start_auto_button.grid(row=9, column=1, padx=10, pady=10)

        create_excel_button = tk.Button(self.auto_sub_frame, text="Escolher Diretório", command=self.choose_directory_excel)
        create_excel_button.grid(row=9, column=0, padx=5, pady=5)

        create_save_button = tk.Button(self.auto_sub_frame, text="Guardar Dados", command=self.save_auto_data_excel)
        create_save_button.grid(row=10, column=0, padx=5, pady=5)

        self.auto_text2 = tk.Text(self.auto_sub_frame, height=10, width=40, wrap=tk.WORD)
        self.auto_text2.grid(row=8, column=1, padx=10, pady=10)
        self.auto_text2.insert("1.0", "F1 23,3 M1 -12,4\nF2 -12,2 M2 10\nF3 -12,3 M3 14\nrep 1\nExemplo de resultados a obter do sistema.")

        create_export_button = tk.Button(self.auto_sub_frame, text="Guardar Resultados", command=self.save_auto_data2_excel)
        create_export_button.grid(row=10, column=1, padx=5, pady=5)

    def tstep(self):
        # Reads the value of the TStep field and converts it to int
        tstep_value = int(self.tstep_entry.get())*0.001
        return tstep_value
    
    def start_auto_control(self):
        self.automatic_mode = True
        mode = "AUTOMODE:ON"
        self.stepper_serial_port.write(f"{mode}\n".encode())
        print(f"Auto {mode}")

        tstep_value = self.tstep()

        rep_value = int(self.rep_entry.get())  # Converts the value to an integer
        tmeas_value = int(self.tmeas_entry.get())
        ncount_value = int(self.ncount_entry.get())
        # Send the rep, tmeas, and ncount values to the Arduino
        #self.stepper_serial_port.write(f"REP:{rep_value}\n".encode())
        self.stepper_serial_port.write(f"TMEAS:{tmeas_value}\n".encode())
        self.stepper_serial_port.write(f"NCOUNT:{ncount_value}\n".encode())
        print(f"rep_value: {rep_value}, tmeas_value: {tmeas_value}, ncount_value: {ncount_value}")
        time.sleep(tstep_value)


        auto_text = self.auto_text.get("1.0", tk.END)
        lines = auto_text.split("\n")
        formatted_responses = []  # To store formatted responses
        current_rep = None  # To track the current replay

        force_values = []# List for storing force values
        moment_values = [] # List for storing momentum values
        ma_valuess = [] # Initialize the list of MA values
        ml_valuess = [] # Initialize the list of ML values

        for _ in range(rep_value):
                if current_rep is not None:
                    # Add a line indicating the repeat change
                    formatted_responses.append(f"Repetição {current_rep + 2}")
                for line in lines:
                    if line.strip():
                        # Send the current line to the Arduino
                        self.stepper_serial_port.write(line.encode())
                        print("Enviando:", line)

                                # Use Regular Expressions to Extract MA and ML Values
                        ma_match = re.search(r"MA (-?\d+(?:\.\d+)?(?:,\d+)?)", line)
                        ml_match = re.search(r"ML (-?\d+(?:\.\d+)?(?:,\d+)?)", line)

                        ma_value = ma_match.group(1) if ma_match else None
                        ml_value = ml_match.group(1) if ml_match else None

                        # Add the MA and ML values to the respective lists
                        ma_valuess.append(ma_value)
                        ml_valuess.append(ml_value)
        
                        # Allow a short time for the Arduino to process the command (adjust as needed)
                        time.sleep(tstep_value)

                        # Read the Arduino response
                        response = self.stepper_serial_port.readline().decode().strip()
                        print("Resposta do Arduino:", response)

                        # Use Regular Expressions to Extract FX and MZ Values
                        fx_match = re.search(r"FX:\s*([\d.]+)", response)
                        mz_match = re.search(r"MZ:\s*([\d.]+)", response)
                        
                        fx_value = fx_match.group(1) if fx_match else None
                        mz_value = mz_match.group(1) if mz_match else None
                        
                        formatted_responses.append(f"Fx {fx_value} Mz {mz_value}")

                        # Add the force and moment values to the lists
                        force_values.append(fx_value)
                        moment_values.append(mz_value)

                # Update the current replay
                current_rep = _

        # Concatenate all responses into a single string
        formatted_text = "\n".join(formatted_responses)

        # Enter the formatted string in the text box auto_text2
        self.auto_text2.delete("1.0", tk.END)
        self.auto_text2.insert(tk.END, formatted_text)

        # Convert string lists to float lists
        force_values = [float(value.replace(',', '.')) for value in force_values]
        moment_values = [float(value.replace(',', '.')) for value in moment_values]
        ml_valuess = [float(value.replace(',', '.')) for value in ml_valuess]
        ma_valuess = [float(value.replace(',', '.')) for value in ma_valuess]
                
        self.plot_graph(force_values, moment_values, ma_valuess, ml_valuess)

    def choose_directory_excel(self):
        # Opens the directory choice window
        self.chosen_directory = filedialog.askdirectory()
        if self.chosen_directory:
            # Here you can save your chosen directory for later use
            print("Diretório escolhido:", self.chosen_directory)

            # Prompts the user to enter the file name
            filename_input = simpledialog.askstring("Nome do arquivo", "Insira o nome do arquivo:")
            if filename_input:
            # Adds the .xlsx extension to the file name if it is not present
                if not filename_input.endswith(".xlsx"):
                    filename_input += ".xlsx"
            self.chosen_filename = filename_input

    def save_auto_data_excel(self):
        if self.chosen_directory:
            filename = f"{self.chosen_directory}/{self.chosen_filename}"
            auto_text = self.auto_text.get("1.0", tk.END).strip()
            auto_text_lines = auto_text.split("\n")

            try:
                wb = openpyxl.load_workbook(filename)
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            ws = wb.active
            if "dados_auto" not in wb.sheetnames:
                ws.title = "dados_auto"
                ws.cell(row=1, column=1, value="Dados de Motores")
                ws.cell(row=2, column=1, value="MA")
                ws.cell(row=2, column=2, value="ML")

            for line in auto_text_lines:
                # Use Regular Expressions to Extract MA and ML Values
                ma_match = re.search(r"MA (-?\d+\.\d+|-?\d+,\d+|-?\d+)", line)
                ml_match = re.search(r"ML (-?\d+\.\d+|-?\d+,\d+|-?\d+)", line)

                ma_value = ma_match.group(1) if ma_match else None
                ml_value = ml_match.group(1) if ml_match else None

                max_row = ws.max_row
                new_row_idx = max_row + 1

                ws.cell(row=new_row_idx, column=1, value=ma_value)
                ws.cell(row=new_row_idx, column=2, value=ml_value)

            wb.save(filename)
            print("Dados automáticos adicionados e salvos em:", filename)
        else:
            print("Escolha um diretório antes de salvar os dados automáticos.")

    def save_auto_data2_excel(self):
        if self.chosen_directory:
            filename = f"{self.chosen_directory}/{self.chosen_filename}"
            auto_text2 = self.auto_text2.get("1.0", tk.END).strip()
            auto_text2_lines = auto_text2.split("\n")
            linhas = auto_text2.count('\n') + 1  # Count the number of new rows and add 1


            try:
                wb = openpyxl.load_workbook(filename)
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            ws = wb.active
            ws.cell(row=1, column=3, value="Dados de Load Cells")
            ws.cell(row=2, column=3, value="Força")
            ws.cell(row=2, column=4, value="Momento")

            max_row = ws.max_row
            new_row_idx = 3
            columnn_r=3

            for line in auto_text2_lines:
                repetition_match = re.search(r"Repetição (\d+)", line)
                
                if repetition_match:
                    repetition_number = int(repetition_match.group(1))
                    columnn_r += 2
                    new_row_idx = 2
                
                fx_match = re.search(r"Fx (-?\d+\.\d+)", line)
                mz_match = re.search(r"Mz (-?\d+\.\d+)", line)

                fx_value = fx_match.group(1) if fx_match else None
                mz_value = mz_match.group(1) if mz_match else None

                ws.cell(row=2, column=columnn_r, value="Força")
                ws.cell(row=2, column=columnn_r+1, value="Momento")
                ws.cell(row=new_row_idx,column=columnn_r, value=fx_value)
                ws.cell(row=new_row_idx,column=columnn_r+1, value=mz_value)

                new_row_idx += 1

            wb.save(filename)
            print("Dados automáticos adicionados e salvos em:", filename)
        else:
            print("Escolha um diretório antes de salvar os dados automáticos.")


    def close(self):
        self.stepper_serial_port.close()
        self.root.destroy()
        
    def send_stepper1_values(self):
        self.automatic_mode = False
        tstep_value = self.tstep()

        # Set the default values for rep, tmeas, and ncount
        rep_value = 1
        tmeas_value = int(self.tmeas_entry.get())
        ncount_value = int(self.ncount_entry.get())

        # Send the rep, tmeas, and ncount values to the Arduino
        #self.stepper_serial_port.write(f"REP:{rep_value}\n".encode())
        self.stepper_serial_port.write(f"TMEAS:{tmeas_value}\n".encode())
        self.stepper_serial_port.write(f"NCOUNT:{ncount_value}\n".encode())
        print(f"rep_value: {rep_value}, tmeas_value: {tmeas_value}, ncount_value: {ncount_value}")
        time.sleep(tstep_value)

        stepper1_value = self.stepper1_entry.get()
        # Send the stepper motor values to the Arduino
        stepper_message = f"MOTOR_LINEAR:{stepper1_value}\n"  # Add ":" as a separator
        self.stepper_serial_port.write(stepper_message.encode())
        print(stepper_message)

        time.sleep(tstep_value)

        # Read the Arduino response
        response = self.stepper_serial_port.readline().decode().strip()
        if response.startswith("FORCA:"):
            forca_value = response[6:]
            self.output_text.delete("1.0", tk.END)
            self.output_text.insert(tk.END, f"{forca_value}")


    def send_stepper2_values(self):
        self.automatic_mode = False
        tstep_value = self.tstep()

        # Set the default values for rep, tmeas, and ncount
        rep_value = 1
        tmeas_value = int(self.tmeas_entry.get())
        ncount_value = int(self.ncount_entry.get())

        # Send the rep, tmeas, and ncount values to the Arduino
        #self.stepper_serial_port.write(f"REP:{rep_value}\n".encode())
        self.stepper_serial_port.write(f"TMEAS:{tmeas_value}\n".encode())
        self.stepper_serial_port.write(f"NCOUNT:{ncount_value}\n".encode())
        print(f"rep_value: {rep_value}, tmeas_value: {tmeas_value}, ncount_value: {ncount_value}")
        time.sleep(tstep_value)

        stepper2_value = self.stepper2_entry.get()
        # Send the stepper motor values to the Arduino
        stepper_message2 = f"MOTOR_ANGULAR:{stepper2_value}\n"  # Adicione ":" como separador
        self.stepper_serial_port.write(stepper_message2.encode())
        print(stepper_message2)

        time.sleep(tstep_value)

        # Read the Arduino response
        response = self.stepper_serial_port.readline().decode().strip()
        if response.startswith("MOMENTO:"):
            momento_value = response[8:]
            self.output_text2.delete("1.0", tk.END)
            self.output_text2.insert(tk.END, f"{momento_value}")

    def close(self):
        self.stepper_serial_port.close()
        self.root.destroy()


    def choose_directory(self):
        # Opens the directory choice window
        self.chosen_directory = filedialog.askdirectory()
        if self.chosen_directory:
            # Saves chosen directory for later use
            print("Diretório escolhido:", self.chosen_directory)

            # Prompts the user to enter the file name
            filename_input = simpledialog.askstring("Nome do arquivo", "Insira o nome do arquivo:")
            if filename_input:
            # Adds the .xlsx extension to the file name if it is not present
                if not filename_input.endswith(".xlsx"):
                    filename_input += ".xlsx"
            self.chosen_filename = filename_input

    def save_data(self):
        print(self.chosen_directory)
        if self.chosen_directory:
            filename2 = f"{self.chosen_directory}/{self.chosen_filename}"
            stepper1_value = self.stepper1_entry.get()
            print(stepper1_value)
            stepper2_value = self.stepper2_entry.get()
            print(stepper2_value)
            output_value = self.output_text.get("1.0", tk.END)
            print(output_value)
            output2_value = self.output_text2.get("1.0", tk.END)
            print(output2_value)

            # Open the existing Excel file
            try:
                wb2 = openpyxl.load_workbook(filename2)
                print('não se criou')
            except FileNotFoundError:
                # If the file does not exist, create a new one
                wb2 = openpyxl.Workbook()
                print('criou')

            ws2 = wb2.active
            max_column = ws2.max_column  # Set max_column

            if "dados_manual" not in wb2.sheetnames:
                ws2.title = "dados_manual"
                ws2.cell(row=1, column=1, value="Dados de Motor Linear (mm)")
                ws2.cell(row=1, column=2, value=stepper1_value)
                ws2.cell(row=2, column=1, value="Dados do Motor Angular (º)")
                ws2.cell(row=2, column=2, value=stepper2_value)
                ws2.cell(row=3, column=1, value="Força de End-Loop Alfa (N)")
                ws2.cell(row=3, column=2, value=output_value)
                ws2.cell(row=4, column=1, value="Momento de End-Loop Beta (N.m)")
                ws2.cell(row=4, column=2, value=output2_value)
            else:
                ws2.cell(row=1, column=max_column + 1, value=stepper1_value)
                ws2.cell(row=2, column=max_column + 1, value=stepper2_value)
                ws2.cell(row=3, column=max_column + 1, value=output_value)
                ws2.cell(row=4, column=max_column + 1, value=output2_value)

            wb2.save(filename2)
            print("Dados adicionados e salvos em:", filename2)
        
        else:
            print("Escolha um diretório antes de salvar os dados.")


    def plot_graph(self,force_values, moment_values, ma_valuess, ml_valuess):
        
        # Create a new plot
        plt.figure(figsize=(8, 6))

        # Plote Força por ML
        plt.subplot(2, 1, 1)  # 2 rows, 1 column, first plot
        plt.plot(ml_valuess, force_values, 'b-', label='Força por ML')
        plt.xlabel('ML')
        plt.ylabel('Força')
        plt.legend()

        # Plote Momento por MA
        plt.subplot(2, 1, 2)  # 2 rows, 1 column, second plot
        plt.plot(ma_valuess, moment_values, 'r-', label='Momento por MA')
        plt.xlabel('MA')
        plt.ylabel('Momento')
        plt.legend()

        # Display the plot in the interface
        plt.tight_layout()
        plt.show()
    
    def move_to_initial1(self):
        # Send command to Arduino to move the Linear Motor to the home position
        # Replace "COMANDO_MOTOR_LINEAR" with the specific command you send to the Arduino
        command = "COMANDO_MOTOR_LINEAR"
        self.stepper_serial_port.write(command.encode())
        print("Movendo Motor Linear para a posição inicial...")

    def move_to_initial2(self):
        # Send command to Arduino to move the Angular Motor to the home position
        # Replace "COMANDO_MOTOR_ANGULAR" with the specific command you send to the Arduino
        command = "COMANDO_MOTOR_ANGULAR"
        self.stepper_serial_port.write(command.encode())
        print("Movendo Motor Angular para a posição inicial...")

    def start_serial_read(self):
        # Create a thread to read the serial data
        self.serial_thread = threading.Thread(target=self.process_serial_input)
        self.serial_thread.daemon = True
        self.serial_thread.start()

    def close(self):
        self.stepper_serial_port.close()

    
if __name__ == "__main__":
    root = tk.Tk()
    app = StepperControlGUI(root)
    root.mainloop()
